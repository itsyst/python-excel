import openpyxl

workbook = openpyxl.load_workbook("orders.xlsx")
print(workbook.sheetnames)
feuil = workbook["Feuil"]
sheet = workbook["Sheet"]

# print("rows:", sheet.max_row)
# print("columns:", sheet.max_column)


# access a cell
# cell = sheet["A1"]
# column = sheet["a"] ## return a tuple of all the cells in the A column
# cells = sheet["a:c"]  # returns a tuple of tuple each tuple represent a column
# cells = sheet["a1:c4"] ## use coordinates
# cells = sheet[1] ## return cells in the first row
# add a row in the the end of the sheet
# feuil.append([10003, 4,  '$20.99'])
# sheet.insert_rows(0,1) ## insert a row in the given index
# sheet.delete_rows(13, 1)
# print(cells)
# cell = sheet.cell(row=1, column=1)
# print(cell.value)
# print(cell.row)
# print(cell.column)
# print("coordinate:", cell.coordinate)

# # get all rows and columns
# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, column)
#         print(cell.value)

workbook.create_sheet("Ark")

# ! command Query separation principle
for row in range(1, 10):
    cell = sheet.cell(row, 3)
    print(cell.value)


workbook.save("_orders.xlsx")
